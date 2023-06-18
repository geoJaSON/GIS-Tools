#!/usr/bin/env python
# coding: utf-8

# In[1]:


print('Getting things ready')
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook
import pandas as pd
import openpyxl
from arcgis.gis import GIS
from datetime import date, timedelta, timezone

#Connect to main layer (SWG PRB v2/Information)

tokens = pd.read_csv(r"C:\Users\M3ECHJJJ\Documents\Tokens.csv", index_col='Site')
gis = GIS(tokens.at['uCOP','URL'])
prb_layer_item = gis.content.get('ab6017089de34b84ac312a051e53af47')
prb_layers = prb_layer_item.layers
prb_fset=prb_layers[3].query()
prb_features = prb_fset.features
prb_flayer = prb_layers[3]
prb_flayer

#Connect to funding related table (SWG PRB v2/Funding)

funding_layers = prb_layer_item.tables
funding_fset=funding_layers[0].query()
funding_features = funding_fset.features
funding_flayer = funding_layers[0]
funding_flayer

#Connect to milestone related table (SWG PRB v2/Milestones)

activities_layers = prb_layer_item.tables
activities_fset=activities_layers[2].query()
activities_features = activities_fset.features
activities_flayer = activities_layers[2]
activities_flayer

#Create 3 pandas dataframe templates to hold project, milestone, and activities informaiton

df = pd.DataFrame(columns = ['project_name','fd_fed_cost','fd_fed_rec_to_date','fd_fed_bal','fd_nf_cost','fd_nf_rec_to_date','fd_nf_bal','fd_total_cost','fd_total_rec_to_date','fd_total_bal','category','district'])
df.set_index("project_name", inplace = True)
schdftemplate = pd.DataFrame(columns = ['project_name','milestone_code','activity_description','baseline_date','basic_date','current_date_','actual_date','status_letter','status_color','parentglobalid'])
schdftemplate.set_index("activity_description", inplace = True)
fundftemplate = pd.DataFrame(columns = ['project_name','funding_type','carryover_from_last_fy','sch_to_receive_current_fy','avail_to_obl_ytd','basic_sched_ytd','actuals_to_date','execution','current_sched_ytd','parentglobalid'])
fundftemplate.set_index("funding_type", inplace = True)
fundf = fundftemplate
fundfitr= fundftemplate
schdf = schdftemplate
schdfitr= schdftemplate


#Convert cell address to usable row, col

def get_col_number(cell):
    xy = coordinate_from_string(cell) # returns ('A',4)
    icol = column_index_from_string(xy[0]) # returns 1
    irow = xy[1]
    return icol, irow


#Find top-left corner of milestone information

def find_schedule_cg():
    for row in sheet.iter_rows():
        for cell in row:
            try:
                if "Milestone Code" in cell.value:
                    return get_col_number(cell.coordinate)
            except:
                pass


#Find top-left corner of funding information

def find_type_of_funds_cg():
    for row in sheet.iter_rows():
        for cell in row:
            try:
                if "Funding Type" in cell.value:
                    return get_col_number(cell.coordinate)
            except:
                pass


#Find top-left corner of financial information

def find_study_cost_cg():
    for row in sheet.iter_rows():
        for cell in row:
            try:
                if "Project" in cell.value:
                    return get_col_number(cell.coordinate)
            except:
                pass

#Pick out financial data from cells according to their relative position from top-left

def update_study_cost_values(sheet,xlabel):

    study_cost_x,study_cost_y = find_study_cost_cg()
    try:
        df.at[sheet.title, 'fd_fed_cost'] = int(sheet.cell(study_cost_y+1, study_cost_x+1).value)
    except:
        df.at[sheet.title, 'fd_fed_cost'] = 0
    try:
        df.at[sheet.title, 'fd_fed_rec_to_date'] = int(sheet.cell(study_cost_y+2, study_cost_x+1).value)
    except:
        df.at[sheet.title, 'fd_fed_rec_to_date'] = 0
    try:
        df.at[sheet.title, 'fd_fed_bal'] = int(sheet.cell(study_cost_y+3, study_cost_x+1).value)
    except:
        df.at[sheet.title, 'fd_fed_bal'] = 0
    try:
        df.at[sheet.title, 'fd_nf_cost'] = int(sheet.cell(study_cost_y+1, study_cost_x+2).value)
    except:
        df.at[sheet.title, 'fd_nf_cost'] = 0
    try:
        df.at[sheet.title, 'fd_nf_rec_to_date'] = int(sheet.cell(study_cost_y+2, study_cost_x+2).value)
    except:
        df.at[sheet.title, 'fd_nf_rec_to_date'] = 0
    try:
        df.at[sheet.title, 'fd_nf_bal'] = int(sheet.cell(study_cost_y+3, study_cost_x+2).value)
    except:
        df.at[sheet.title, 'fd_nf_bal'] = 0
    try:
        df.at[sheet.title, 'fd_total_cost'] = int(sheet.cell(study_cost_y+1, study_cost_x+3).value)
    except:
        df.at[sheet.title, 'fd_total_cost'] = 0
    try:
        df.at[sheet.title, 'fd_total_rec_to_date'] = int(sheet.cell(study_cost_y+2, study_cost_x+3).value)
    except:
        df.at[sheet.title, 'fd_total_rec_to_date'] = 0
    try:
        df.at[sheet.title, 'fd_total_bal'] = int(sheet.cell(study_cost_y+3, study_cost_x+3).value)
    except:
        df.at[sheet.title, 'fd_total_bal'] = 0
    try:
        df.at[sheet.title, 'category'] = xlabel
    except:
        df.at[sheet.title, 'category'] = ''
    df.at[sheet.title, 'district'] = 'SWG'


#Give network locations of Construction, Investigation, PL-8499, and CAP spreadsheets

xls_list = {r"S:\Shared Files\Project Information Folders\PRB - RPMR Slides\2 - Construction\CG_Financial_Milestone_Tables_for_PRB_slides.xlsx":'Construction',
    r'S:\Shared Files\Project Information Folders\PRB - RPMR Slides\1 - Investigations\GI_Financial_Milestone_Tables_for_PRB_slides.xlsx':'Investigation',
    r"S:\Shared Files\Project Information Folders\PRB - RPMR Slides\PL84-99 Projects\PL84-99_Financial_Milestone_Tables_for_PRB_slides.xlsx":"PL84-99",
    r"S:\Shared Files\Project Information Folders\PRB - RPMR Slides\3 - CAP\CAP_Financial_Milestone_Tables_for_PRB_slides (version 1).xlsb.xlsx":"CAP"}


#Loop through each spreadsheet, append projects dataframe and grab milestone & funding dataframes
print('Getting project data from spreadsheets')
for xls,label in xls_list.items():
    print(xls)
    wb = load_workbook(xls, data_only = True)
    for sheet in wb.worksheets: #loop through projects in spreadsheet
        print('    ',sheet.title)
        study_cost_x,study_cost_y = find_study_cost_cg() #Find top-left corner of project funding table
        update_study_cost_values(sheet,label) #

        schedule_x,schedule_y  = find_schedule_cg()
        i = 1
        schdfitr = schdftemplate
        while True: #iterate rows below top-left of milestone info, stopping when reading the cell with "Bold" in text
            try:
                if 'Bold' not in sheet.cell(schedule_y+i, schedule_x).value and sheet.cell(schedule_y+i, schedule_x).value != None:
                    schdfitr = schdfitr.append(pd.Series(name=sheet.cell(schedule_y+i, schedule_x+1).value))
                    schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'project_name'] = sheet.title
                    schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'milestone_code'] = sheet.cell(schedule_y+i, schedule_x).value
                    schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'baseline_date'] = sheet.cell(schedule_y+i, schedule_x+2).value
                    schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'basic_date'] = sheet.cell(schedule_y+i, schedule_x+3).value
                    schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'current_date_'] = sheet.cell(schedule_y+i, schedule_x+4).value
                    schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'actual_date'] = sheet.cell(schedule_y+i, schedule_x+5).value
                    schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'status_letter'] = sheet.cell(schedule_y+i, schedule_x+6).value
                    try:
                        schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'status_color'] = sheet.cell(schedule_y+i, schedule_x+6).fill.start_color.index[2:]
                    except:
                        schdfitr.at[sheet.cell(schedule_y+i, schedule_x+1).value, 'status_color'] = '000000'
                    i = i + 1
                else:
                    break
            except:
                break
        schdf = schdf.append(schdfitr)
        type_of_funds_x,type_of_funds_y= find_type_of_funds_cg()
        i = 1
        fundfitr = fundftemplate
        while True:
            if  sheet.cell(type_of_funds_y+i, type_of_funds_x).value != None:
                fundfitr = fundfitr.append(pd.Series(name=sheet.cell(type_of_funds_y+i, type_of_funds_x).value))
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'project_name'] = sheet.title
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'carryover_from_last_fy'] = sheet.cell(type_of_funds_y+i, type_of_funds_x+1).value
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'sch_to_receive_current_fy'] = sheet.cell(type_of_funds_y+i, type_of_funds_x+2).value
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'avail_to_obl_ytd'] = sheet.cell(type_of_funds_y+i, type_of_funds_x+3).value
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'basic_sched_ytd'] = sheet.cell(type_of_funds_y+i, type_of_funds_x+4).value
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'actuals_to_date'] = sheet.cell(type_of_funds_y+i, type_of_funds_x+5).value
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'execution'] = sheet.cell(type_of_funds_y+i, type_of_funds_x+6).value
                fundfitr.at[sheet.cell(type_of_funds_y+i, type_of_funds_x).value, 'current_sched_ytd'] = sheet.cell(type_of_funds_y+i, type_of_funds_x+7).value

                i = i + 1
            else:
                break
        fundf = fundf.append(fundfitr)

#Clean up funding and milestone tables
print('Cleaning null values')
fundf.loc[(fundf.execution == '#DIV/0!'),'execution']=0
schdf.loc[(schdf.status_color == '000000'),'status_color']='00'


#%% Sync projects dataframe with Information layer, updates existing projects. Adds new projects if project_key isn't found.
#project_key is title of project tabs in each spreadsheet

def assemble_proj_updates(row):
    if 'Blank' not in row.name and row.name != 'Overview':
        try:
            original_feature = [f for f in prb_features if f.attributes['project_key'] ==row.name][0]
            edit_feature={'objectid':original_feature.attributes['objectid']}
            edit_feature['fed_cost'] = row['fd_fed_cost']
            edit_feature['fed_rec_to_date'] = row['fd_fed_rec_to_date']
            edit_feature['fed_bal'] = row['fd_fed_bal']
            edit_feature['nf_cost'] = row['fd_nf_cost']
            edit_feature['nf_rec_to_date'] = row['fd_nf_rec_to_date']
            edit_feature['nf_bal'] = row['fd_nf_bal']
            edit_feature['total_cost'] = row['fd_total_cost']
            edit_feature['total_rec_to_date'] = row['fd_total_rec_to_date']
            edit_feature['total_bal'] = row['fd_total_bal']
            feat = {'attributes': edit_feature}
            updates_to_push.append(feat)
        except IndexError:
            add_feature = {}
            add_feature['project_name'] = row.name
            add_feature['project_key'] = row.name
            add_feature['project_type'] = row['category']
            add_feature['fed_cost'] = row['fd_fed_cost']
            add_feature['fed_rec_to_date'] = row['fd_fed_rec_to_date']
            add_feature['fed_bal'] = row['fd_fed_bal']
            add_feature['nf_cost'] = row['fd_nf_cost']
            add_feature['nf_rec_to_date'] = row['fd_nf_rec_to_date']
            add_feature['nf_bal'] = row['fd_nf_bal']
            add_feature['total_cost'] = row['fd_total_cost']
            add_feature['total_rec_to_date'] = row['fd_total_rec_to_date']
            add_feature['total_bal'] = row['fd_total_bal']
            add_feature['district'] = row['district']
            add_feature['hidedelete'] = 'No'
            feat = {'attributes': add_feature}
            adds_to_push.append(feat)

updates_to_push = [] #Create list of all update objects to push
adds_to_push = [] #Create list of all insert objects to push
df.apply(lambda row: assemble_proj_updates(row), axis=1) #apply the assumble updates function to DF
update_result = prb_flayer.edit_features(updates=updates_to_push, adds=adds_to_push, rollback_on_failure=False) #Commit the update and insert lists

#Refresh Information layer, getting globalid for each project in order to use as parentglobalid for milestone and funding
prb_fset=prb_layers[3].query()
prb_features = prb_fset.features
prb_flayer = prb_layers[3]
prb_flayer
globalid_dict = {}
for f in prb_features:
    globalid_dict[f.attributes['project_key']] = f.attributes['globalid']


#%% Sync milestone dataframe to feature layer. Attempts to update if milestone has already been pushed, otherwise adds as new record

def assemble_schedule_updates(row):
    try:
        pid=globalid_dict[row['project_name']]
        add_feature = {"parentglobalid": pid}
        add_feature["activity_description"] = row.name
        add_feature["milestone_code"] = row['milestone_code']
        try:
            date_time = row['baseline_date']
            add_feature["baseline_date"]= int(date_time.replace(tzinfo=timezone.utc).timestamp())*1000+63200000
        except:
            pass
        try:
            date_time = row['basic_date']
            add_feature["basic_date"]= int(date_time.replace(tzinfo=timezone.utc).timestamp())*1000+63200000
        except:
            pass
        try:
            date_time = row['current_date_']
            add_feature["current_date_"]= int(date_time.replace(tzinfo=timezone.utc).timestamp())*1000+63200000
        except:
            pass
        try:
            date_time = row['actual_date']
            add_feature["actual_date"] = int(date_time.replace(tzinfo=timezone.utc).timestamp())*1000+63200000
        except:
            pass

        add_feature["status_letter"] = row['status_letter']
        add_feature["status_color"] = row['status_color']
        feat = {'attributes': add_feature}
        adds_to_push.append(feat)
    except Exception as e:
        print(e)

updates_to_push = [] #Create list of all update objects to push
adds_to_push = [] #Create list of all insert objects to push
schdf.apply(lambda row: assemble_schedule_updates(row), axis=1) #apply the assumble updates function to DF
activities_flayer.delete_features(where = "objectid > 0")
update_result = activities_flayer.edit_features(updates=updates_to_push, adds=adds_to_push, rollback_on_failure=False) #Commit the update and insert lists


#%%
def assemble_funding_updates(row):
    try:
        add_feature = {"parentglobalid": globalid_dict[row['project_name']]}
        add_feature["funding_type"] = row.name
        add_feature["carryover_from_last_fy"] = row['carryover_from_last_fy']
        add_feature["sch_to_receive_current_fy"] = row['sch_to_receive_current_fy']
        add_feature["avail_to_obli_ytd"] = row['avail_to_obl_ytd']
        add_feature["basic_sched_ytd"] = row['basic_sched_ytd']
        add_feature["actuals_to_date"] = row['actuals_to_date']
        add_feature["execution"] = 0
        add_feature["current_sched_ytd"] = row['current_sched_ytd']
        feat = {'attributes': add_feature}
        adds_to_push.append(feat)
    except Exception:
        pass

updates_to_push = [] #Create list of all update objects to push
adds_to_push = [] #Create list of all insert objects to push
fundf.apply(lambda row: assemble_funding_updates(row), axis=1) #apply the assumble updates function to DF
funding_flayer.delete_features(where = "objectid > 0")
update_result = funding_flayer.edit_features(updates=updates_to_push, adds=adds_to_push, rollback_on_failure=False) #Commit the update and insert lists



############################################################################

update_result
updates_to_push
adds_to_push
